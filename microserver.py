import openpyxl as opx

sap_number = input('введите sap номер магазина:')

workbook_path = 'c:\Programs\Перекресток_общий_2020.xlsx'
workbook = opx.load_workbook(workbook_path, read_only=True)

first_sheet = workbook['СМ']

first_sheet_max_row = first_sheet.max_row
first_sheet_max_column = first_sheet.max_column

print('max row : ', first_sheet_max_row, 'max col : ', first_sheet_max_column)

for i in range(1, first_sheet_max_row):
    cell = first_sheet.cell(row=i, column=12)
    if cell.value == sap_number:
        print(first_sheet.cell(row=i, column=11).value)
        print(first_sheet.cell(row=i, column=14).value)
